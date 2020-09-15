"""
Created on Sat Sep 12 22:32:41 2020
Generate 200 numbers and find probability of any random digit
@author: Rubeena
"""
import numpy as np
import random
import pandas as pd
import openpyxl

sample_size=200
 
number=np.random.randint(1, 99999999, size=sample_size)
print("random numbers generated are",number)

#Frequency distribution table as given in the problem in Excel sheet
digit=np.array([0,1,2,3,4,5,6,7,8,9])
Frequency=np.array([22,26,22,22,20,10,14,28,16,20])
freqTable=pd.DataFrame({'Digit':digit,'Frequency':Frequency})
filepath_givenDta='Freq_Table_given.xlsx'
freqTable.to_excel(filepath_givenDta,'sheet1',index=False)
d1=pd.Series(Frequency)


def FindFreqTable(unitDigits,digit):          #accordingly check value of frequency of last digit
    freq1=np.zeros((len(digit),),dtype='int')
    for i in range(0,len(unitDigits)):
        for j in range(0,len(digit)):
            if unitDigits[i]==digit[j]:
                c=j
                freq1[c]+=1
    return freq1;

def LastDigitArray(number):
    lastDigit = []
    for i in range(0,len(number)):
        lastDigit.append(number[i]%10)
    return lastDigit;

#Enter Random numbers and their last digits into excel sheet
unitDigits=LastDigitArray(number)
print("Last digits of each of these numbers generated are:\n",unitDigits)
randNums=pd.DataFrame({'Numbers':number,'Unit Digit':unitDigits})
filepath_randNums='Random_numbers_generated.xlsx'
randNums.to_excel(filepath_randNums,'RandomNumbers',index=False)

#This part prints frequency of each digit on console.
#Used only for comparison purpose
frequen=FindFreqTable(unitDigits,digit)
d2=pd.Series(frequen)
print("freq of digits of random num generated is \n",d2)

#Generates an excel sheet containing Frequency Distribution Table for Random numbers unit digits
def FreqOfRandomDigits(digit):
    freq=np.zeros((len(digit),),dtype='int')
    ps_random = openpyxl.load_workbook('Random_numbers_generated.xlsx')
    sheet = ps_random['RandomNumbers']
    sheet.max_row   #returns the total number of rows in the sheet
    for row in range(2, sheet.max_row + 1):
        for j in range(0,len(digit)):
            if sheet['B'+str(row)].value==digit[j]:
                c=j
                freq[c]+=1
    freq_random=pd.DataFrame({"Digit":digit,"Frequency":freq})
    filepath = 'Freq_Table_generated_random.xlsx'
    freq_random.to_excel(filepath,'sheet_random',index=False)
    return;

FreqOfRandomDigits(digit)

#Chose a random last digit
randomLastDigit=random.randint(0,len(digit)-1)
print("\n A random last digit chosen: ", randomLastDigit)

#Find probability of a randomly selected unit digit
def FindProbability(randomLastDigit):
    ps_random = openpyxl.load_workbook('Freq_Table_generated_random.xlsx')
    sheet_freq = ps_random['sheet_random']
    sheet_freq.max_row   #returns the total number of rows in the sheet
    for row in range(2, sheet_freq.max_row + 1):
        if sheet_freq['A'+str(row)].value==randomLastDigit:
            probability=(sheet_freq['A'+str(row)].value)/sample_size
    print("Probability of randomly chosen digit",randomLastDigit,"as unit digit is: ",probability)
    return;

FindProbability(randomLastDigit)
